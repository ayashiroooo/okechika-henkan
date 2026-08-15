from openpyxl import load_workbook
import json
from pathlib import Path

# =========================
# 設定
# =========================

# この sync.py と同じフォルダにExcelを置く想定
EXCEL_FILE = "桶文字対応表 from MAKU.xlsx"
SHEET_NAME = "対応表"

RULES_FILE = "rules.json"
EXAMPLES_FILE = "examples.json"

# Excelの列
BEFORE_COLUMN = 2  # B列：変換前
AFTER_COLUMN = 3   # C列：変換後
EXAMPLE_COLUMN = 4 # D列：用例
SOURCE_START_COLUMN = 5  # E列以降：出典

# データ開始行
START_ROW = 3


# =========================
# Excel → JSON
# =========================

def main():

    base_path = Path(__file__).parent

    excel_path = base_path / EXCEL_FILE
    rules_path = base_path / RULES_FILE
    examples_path = base_path / EXAMPLES_FILE


    # =========================
    # Excel確認
    # =========================

    if not excel_path.exists():
        print(f"Excelファイルが見つかりません: {excel_path}")
        return


    wb = load_workbook(
        excel_path,
        data_only=True
    )


    if SHEET_NAME not in wb.sheetnames:
        print(f"シート「{SHEET_NAME}」が見つかりません。")
        print(
            f"利用可能なシート: {', '.join(wb.sheetnames)}"
        )
        return


    ws = wb[SHEET_NAME]


    # =========================
    # rules.json
    # =========================

    rules = {}


    # =========================
    # examples.json
    # =========================

    examples = {}


    # =========================
    # Excelを読む
    # =========================

    for row in range(
        START_ROW,
        ws.max_row + 1
    ):

        before = ws.cell(
            row=row,
            column=BEFORE_COLUMN
        ).value

        after = ws.cell(
            row=row,
            column=AFTER_COLUMN
        ).value

        usage = ws.cell(
            row=row,
            column=EXAMPLE_COLUMN
        ).value


        # ---------------------------------
        # B列またはC列が空欄ならスキップ
        # ---------------------------------

        if before is None or after is None:
            continue


        before = str(before)
        after = str(after)


        # =================================
        # rules.json
        # =================================

        rules[before] = after


        # =================================
        # examples.json
        # =================================
        #
        # 「?」付きのものだけ登録
        #
        # 例：
        #
        # 侚 → 越?
        #
        # =================================

        if not after.endswith("?"):
            continue


        # 新しい項目を作成

        examples[before] = {
            "translation": after,
            "examples": [],
            "sources": []
        }


        # ---------------------------------
        # D列：用例
        # ---------------------------------

        if (
            usage is not None
            and str(usage).strip()
        ):

            examples[before]["examples"].append(
                str(usage)
            )


        # ---------------------------------
        # E列以降：出典
        # ---------------------------------

        for column in range(
            SOURCE_START_COLUMN,
            ws.max_column + 1
        ):

            source = ws.cell(
                row=row,
                column=column
            ).value


            if (
                source is not None
                and str(source).strip()
            ):

                examples[before]["sources"].append(
                    str(source)
                )


    # =========================
    # rules.jsonを書き出す
    # =========================

    with open(
        rules_path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            rules,
            f,
            ensure_ascii=False,
            indent=4
        )


    # =========================
    # examples.jsonを書き出す
    # =========================

    with open(
        examples_path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            examples,
            f,
            ensure_ascii=False,
            indent=4
        )


    # =========================
    # 結果表示
    # =========================

    print("================================")
    print("JSONを更新しました！")
    print("================================")

    print()
    print(
        f"rules.json 登録件数: "
        f"{len(rules)}"
    )

    print(
        f"examples.json 登録件数: "
        f"{len(examples)}"
    )

    print()

    print(
        f"rules.json : {rules_path}"
    )

    print(
        f"examples.json : {examples_path}"
    )

    print("================================")


if __name__ == "__main__":
    main()